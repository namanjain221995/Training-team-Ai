import io
import re
import time
import json
import tempfile
import os
import random
from pathlib import Path
from typing import Optional, Dict, List, Tuple, Any

from dotenv import load_dotenv

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload, MediaFileUpload
from googleapiclient.errors import HttpError
from google.oauth2 import service_account

from openai import OpenAI

load_dotenv()

SLOT_CHOICE = (os.getenv("SLOT_CHOICE") or "").strip()
DEFAULT_OPENAI_MODEL = (os.getenv("OPENAI_MODEL") or "").strip() or "gpt-5.2-mini"

AUTH_MODE = (os.getenv("AUTH_MODE") or "service_account").strip().lower()
SERVICE_ACCOUNT_FILE = Path((os.getenv("SERVICE_ACCOUNT_FILE") or "service-account.json").strip())
USE_DELEGATION = (os.getenv("USE_DELEGATION") or "").strip().lower() in ("1", "true", "yes", "y")
DELEGATED_USER_EMAIL = (os.getenv("DELEGATED_USER_EMAIL") or "").strip()

SHARED_DRIVE_NAME = (os.getenv("SHARED_DRIVE_NAME") or "").strip()
ROOT_2026_FOLDER_NAME = (os.getenv("ROOT_2026_FOLDER_NAME") or "2026").strip()
OUTPUT_ROOT_FOLDER_NAME = (os.getenv("OUTPUT_ROOT_FOLDER_NAME") or "Candidate Result").strip()

ROOT_2026_FOLDER_ID = (os.getenv("ROOT_2026_FOLDER_ID") or "").strip()
OUTPUT_ROOT_FOLDER_ID = (os.getenv("OUTPUT_ROOT_FOLDER_ID") or "").strip()

SCOPES = ["https://www.googleapis.com/auth/drive"]

FOLDER_MIME = "application/vnd.google-apps.folder"
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

FOLDER_NAMES = [
    "3. Introduction Video",
    "4. Mock Interview (First Call)",
    "5. Project Scenarios",
    "6. 30 Questions Related to Their Niche",
    "7. 50 Questions Related to the Resume",
    "8. Tools & Technology Videos",
    "9. System Design Video (with Draw.io)",
    "10. Persona Video",
    "11. Small Talk",
    "12. JD Video",
]

AVG_HEADER = "Average % (available)"
OUTPUT_XLSX_NAME = "Deliverables Analysis Sheet.xlsx"

SKIP_PERSON_FOLDERS = {
    x.strip() for x in (os.getenv("SKIP_PERSON_FOLDERS") or "1. Format").split(",") if x.strip()
}
SKIP_DELIVERABLE_FOLDERS = {
    x.strip() for x in (os.getenv("SKIP_DELIVERABLE_FOLDERS") or "1. Format").split(",") if x.strip()
}

LLM_OUTPUT_FILE_RE = re.compile(r"^LLM_OUTPUT__.*(\.txt)?$", re.I)

FAIL_PASS_PCT_RX = re.compile(
    r"final\s*overall\s*score\s*[:\-]\s*(pass|fail)\s*(?:\(\s*([0-9]+(?:\.[0-9]+)?)\s*%\s*\)|[:\-]?\s*([0-9]+(?:\.[0-9]+)?)\s*%?)",
    re.I,
)

SCORE_REGEXES = [
    re.compile(r"final\s*overall\s*score\s*[:\-]\s*([0-9]+(?:\.[0-9]+)?)\s*(/10|\/\s*10|out\s*of\s*10|%)?", re.I),
    re.compile(r"final\s*score\s*[:\-]\s*([0-9]+(?:\.[0-9]+)?)\s*(/10|\/\s*10|out\s*of\s*10|%)?", re.I),
]

MAX_CHARS_TO_MODEL = 80_000
PCT_MIN = 0.0
PCT_MAX = 100.0
SLEEP_BETWEEN_UPLOADS_SEC = 0.25

EDITOR_EMAILS = [
    "rajvi.patel@techsarasolutions.com",
    "sahil.patel@techsarasolutions.com",
    "soham.piprotar@techsarasolutions.com",
]

THRESHOLDS_PCT: Dict[str, float] = {
    "3. Introduction Video": 75.0,
    "4. Mock Interview (First Call)": 90.0,
    "5. Project Scenarios": 80.0,
    "6. 30 Questions Related to Their Niche": 60.0,
    "7. 50 Questions Related to the Resume": 80.0,
    "8. Tools & Technology Videos": 70.0,
    "9. System Design Video (with Draw.io)": 80.0,
    "10. Persona Video": 70.0,
    "11. Small Talk": 75.0,
    "12. JD Video": 75.0,
    AVG_HEADER: 75.0,
}

GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

RETRYABLE_STATUS_CODES = {429, 500, 502, 503, 504}
MAX_RETRIES = 6

_SLOT_PREFIX_RE = re.compile(r"^\s*(\d+)\.\s*")


def drive_execute(req, label: str = "Drive API call"):
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            return req.execute()
        except HttpError as e:
            status = getattr(e.resp, "status", None)
            if status in RETRYABLE_STATUS_CODES and attempt < MAX_RETRIES:
                sleep_s = min(2 ** (attempt - 1), 20) + random.uniform(0, 1)
                print(
                    f"[RETRY] {label} failed with HTTP {status}. "
                    f"Attempt {attempt}/{MAX_RETRIES}. Sleeping {sleep_s:.1f}s..."
                )
                time.sleep(sleep_s)
                continue
            raise
        except Exception as e:
            if attempt < MAX_RETRIES:
                sleep_s = min(2 ** (attempt - 1), 20) + random.uniform(0, 1)
                print(
                    f"[RETRY] {label} failed ({e}). "
                    f"Attempt {attempt}/{MAX_RETRIES}. Sleeping {sleep_s:.1f}s..."
                )
                time.sleep(sleep_s)
                continue
            raise


def init_openai_client() -> OpenAI:
    api_key = (os.getenv("OPENAI_API_KEY") or "").strip()
    if not api_key:
        raise RuntimeError("OPENAI_API_KEY not found in .env or environment variables.")
    return OpenAI(api_key=api_key)


def safe_response_output_text(resp) -> str:
    if hasattr(resp, "output_text") and resp.output_text:
        return resp.output_text

    out = getattr(resp, "output", None) or []
    parts = []
    for item in out:
        for c in (getattr(item, "content", None) or []):
            t = getattr(c, "text", None)
            if t:
                parts.append(t)
    return "\n".join(parts).strip()


def get_drive_service():
    if AUTH_MODE != "service_account":
        raise RuntimeError(f"Unsupported AUTH_MODE='{AUTH_MODE}'. Use AUTH_MODE=service_account")

    if not SERVICE_ACCOUNT_FILE.exists():
        raise FileNotFoundError(
            f"Service account file not found: {SERVICE_ACCOUNT_FILE}\n"
            f"Set SERVICE_ACCOUNT_FILE in .env or place service-account.json next to this script."
        )

    creds = service_account.Credentials.from_service_account_file(
        str(SERVICE_ACCOUNT_FILE),
        scopes=SCOPES,
    )

    if USE_DELEGATION:
        if not DELEGATED_USER_EMAIL:
            raise RuntimeError("USE_DELEGATION=true but DELEGATED_USER_EMAIL is empty.")
        creds = creds.with_subject(DELEGATED_USER_EMAIL)
        print(f"[AUTH] Service account with delegation -> {DELEGATED_USER_EMAIL}")
    else:
        print("[AUTH] Service account without delegation")

    return build("drive", "v3", credentials=creds)


def _list_kwargs_for_drive(drive_id: Optional[str] = None) -> Dict[str, Any]:
    kwargs: Dict[str, Any] = {
        "supportsAllDrives": True,
        "includeItemsFromAllDrives": True,
    }
    if drive_id:
        kwargs["corpora"] = "drive"
        kwargs["driveId"] = drive_id
    else:
        kwargs["corpora"] = "allDrives"
    return kwargs


def _kwargs_for_mutation() -> Dict[str, Any]:
    return {"supportsAllDrives": True}


def _get_media_kwargs() -> Dict[str, Any]:
    return {"supportsAllDrives": True}


def _escape_drive_q_value(s: str) -> str:
    return s.replace("\\", "\\\\").replace("'", "\\'")


def extract_slot_prefix(name: str) -> Optional[int]:
    m = _SLOT_PREFIX_RE.match(name or "")
    return int(m.group(1)) if m else None


def list_shared_drives(service) -> List[dict]:
    out = []
    page_token = None
    while True:
        res = drive_execute(
            service.drives().list(
                pageSize=100,
                pageToken=page_token,
            ),
            label="list shared drives",
        )
        out.extend(res.get("drives", []))
        page_token = res.get("nextPageToken")
        if not page_token:
            break
    return out


def get_shared_drive_by_name(service, drive_name: str) -> dict:
    drives = list_shared_drives(service)
    matches = [d for d in drives if d.get("name") == drive_name]

    if not matches:
        available = sorted([d.get("name", "") for d in drives])
        raise RuntimeError(
            f"Shared Drive '{drive_name}' not found.\n"
            f"Visible Shared Drives: {available}"
        )

    if len(matches) > 1:
        print(f"[WARN] Multiple Shared Drives named '{drive_name}'. Using first match.")

    return matches[0]


def drive_get_file(service, file_id: str) -> dict:
    return drive_execute(
        service.files().get(
            fileId=file_id,
            fields="id,name,parents,modifiedTime,mimeType,driveId",
            **_kwargs_for_mutation(),
        ),
        label=f"get file {file_id}",
    )


def drive_list_children(service, parent_id: str, mime_type: Optional[str] = None, drive_id: Optional[str] = None):
    q_parts = [f"'{parent_id}' in parents", "trashed = false"]
    if mime_type:
        q_parts.append(f"mimeType = '{mime_type}'")
    q = " and ".join(q_parts)

    page_token = None
    while True:
        res = drive_execute(
            service.files().list(
                q=q,
                fields="nextPageToken, files(id,name,mimeType,modifiedTime,size,parents,driveId)",
                pageSize=1000,
                pageToken=page_token,
                **_list_kwargs_for_drive(drive_id),
            ),
            label=f"list children of parent {parent_id}",
        )
        for f in res.get("files", []):
            yield f
        page_token = res.get("nextPageToken")
        if not page_token:
            break


def drive_find_child(
    service,
    parent_id: str,
    name: str,
    mime_type: Optional[str] = None,
    drive_id: Optional[str] = None,
) -> Optional[dict]:
    safe = _escape_drive_q_value(name)
    q_parts = [f"'{parent_id}' in parents", "trashed = false", f"name = '{safe}'"]
    if mime_type:
        q_parts.append(f"mimeType = '{mime_type}'")
    q = " and ".join(q_parts)

    res = drive_execute(
        service.files().list(
            q=q,
            fields="files(id,name,mimeType,parents,modifiedTime,driveId)",
            pageSize=50,
            **_list_kwargs_for_drive(drive_id),
        ),
        label=f"find child '{name}' in parent {parent_id}",
    )
    files = res.get("files", []) or []
    if not files:
        return None
    return sorted(files, key=lambda f: f.get("modifiedTime") or "", reverse=True)[0]


def drive_create_folder(service, parent_id: str, name: str) -> str:
    meta = {"name": name, "mimeType": FOLDER_MIME, "parents": [parent_id]}
    created = drive_execute(
        service.files().create(
            body=meta,
            fields="id",
            **_kwargs_for_mutation(),
        ),
        label=f"create folder '{name}'",
    )
    return created["id"]


def drive_download_text(service, file_id: str, mime_type: Optional[str]) -> str:
    fh = io.BytesIO()
    try:
        if mime_type and mime_type.startswith("application/vnd.google-apps."):
            request = service.files().export_media(fileId=file_id, mimeType="text/plain")
        else:
            request = service.files().get_media(fileId=file_id, **_get_media_kwargs())

        downloader = MediaIoBaseDownload(fh, request, chunksize=1024 * 1024 * 4)
        done = False
        while not done:
            _, done = downloader.next_chunk()
        fh.seek(0)
        raw = fh.read()
    except HttpError as e:
        print(f"     [WARN] Could not download fileId={file_id}: {e}")
        return ""

    try:
        return raw.decode("utf-8")
    except Exception:
        return raw.decode("latin-1", errors="ignore")


def drive_upload_xlsx(service, parent_id: str, filename: str, local_path: Path, drive_id: Optional[str] = None) -> str:
    matches = [
        f for f in drive_list_children(service, parent_id, None, drive_id)
        if f.get("name") == filename and f.get("mimeType") != FOLDER_MIME
    ]
    existing = None
    if matches:
        existing = sorted(matches, key=lambda x: (x.get("modifiedTime") or ""), reverse=True)[0]

    media = MediaFileUpload(str(local_path), mimetype=XLSX_MIME, resumable=True)

    if existing:
        drive_execute(
            service.files().update(
                fileId=existing["id"],
                media_body=media,
                **_kwargs_for_mutation(),
            ),
            label=f"update xlsx '{filename}'",
        )
        return existing["id"]
    else:
        meta = {"name": filename, "parents": [parent_id]}
        created = drive_execute(
            service.files().create(
                body=meta,
                media_body=media,
                fields="id",
                **_kwargs_for_mutation(),
            ),
            label=f"create xlsx '{filename}'",
        )
        return created["id"]


def drive_search_folder_anywhere_in_shared_drive(service, folder_name: str, drive_id: str) -> List[dict]:
    safe_name = _escape_drive_q_value(folder_name)
    q = f"name = '{safe_name}' and mimeType = '{FOLDER_MIME}' and trashed = false"

    out = []
    page_token = None
    while True:
        res = drive_execute(
            service.files().list(
                q=q,
                fields="nextPageToken, files(id,name,parents,modifiedTime,driveId,mimeType)",
                pageSize=1000,
                pageToken=page_token,
                **_list_kwargs_for_drive(drive_id),
            ),
            label=f"search folder '{folder_name}' in shared drive {drive_id}",
        )
        out.extend(res.get("files", []))
        page_token = res.get("nextPageToken")
        if not page_token:
            break
    return out


def pick_best_named_folder(candidates: List[dict]) -> dict:
    return sorted(candidates, key=lambda c: (c.get("modifiedTime") or ""), reverse=True)[0]


def drive_grant_editor_access(service, file_id: str, emails: List[str]):
    for email in emails:
        perm = {"type": "user", "role": "writer", "emailAddress": email}
        try:
            drive_execute(
                service.permissions().create(
                    fileId=file_id,
                    body=perm,
                    sendNotificationEmail=False,
                    **_kwargs_for_mutation(),
                ),
                label=f"grant editor access to {email}",
            )
            print(f"  [PERM] Editor added: {email}")
        except HttpError as e:
            msg = (e.content or b"").decode("utf-8", errors="ignore").lower()
            if "alreadyexists" in msg or "already exists" in msg or "duplicate" in msg:
                print(f"  [PERM] Already has access: {email}")
            else:
                print(f"  [PERM] Failed for {email}: {e}")
        time.sleep(0.08)


def list_slot_folders(service, slots_parent_id: str, drive_id: str) -> List[dict]:
    folders = list(drive_list_children(service, slots_parent_id, FOLDER_MIME, drive_id))
    out = []

    for f in folders:
        slot_no = extract_slot_prefix(f.get("name", ""))
        if slot_no is not None:
            item = dict(f)
            item["_slot_no"] = slot_no
            out.append(item)

    return sorted(out, key=lambda x: (x["_slot_no"], (x.get("name") or "").lower()))


def choose_slot(service, slots_parent_id: str, drive_id: str) -> dict:
    slots = list_slot_folders(service, slots_parent_id, drive_id)
    if not slots:
        raise RuntimeError(f"No numbered slot folders found under '{ROOT_2026_FOLDER_NAME}' inside the selected Shared Drive.")

    if SLOT_CHOICE.isdigit():
        wanted_slot_no = int(SLOT_CHOICE)
        for s in slots:
            if s["_slot_no"] == wanted_slot_no:
                print(f"[AUTO] Using SLOT_CHOICE={wanted_slot_no}: {s['name']}")
                return s
        available = ", ".join(str(s["_slot_no"]) for s in slots)
        raise RuntimeError(f"SLOT_CHOICE '{wanted_slot_no}' not found. Available folder numbers: {available}")

    print("\n" + "=" * 80)
    print("SELECT SLOT TO PROCESS")
    print("=" * 80)
    for s in slots:
        print(f"  {s['_slot_no']:2}. {s['name']}")
    print("  EXIT - Exit\n")

    while True:
        choice = input("Choose slot number (e.g. 8 or 9) or EXIT: ").strip().lower()
        if choice == "exit":
            raise SystemExit(0)
        if choice.isdigit():
            wanted_slot_no = int(choice)
            for s in slots:
                if s["_slot_no"] == wanted_slot_no:
                    return s
        print("Invalid choice. Try again.")


def clamp_pct(x: Optional[float]) -> Optional[float]:
    if x is None:
        return None
    if x < PCT_MIN or x > PCT_MAX:
        return None
    return x


def normalize_to_percent(value: float, unit: Optional[str]) -> Optional[float]:
    u = (unit or "").lower().strip()

    if "%" in u:
        return value
    if "/10" in u or "out of 10" in u:
        return value * 10.0

    if value <= 10.0:
        return value * 10.0
    if 10.0 < value <= 100.0:
        return value
    return None


def extract_score_regex_percent(text: str) -> Optional[float]:
    if not text:
        return None

    m = FAIL_PASS_PCT_RX.search(text)
    if m:
        pct_str = m.group(2) or m.group(3)
        try:
            return clamp_pct(round(float(pct_str), 2))
        except Exception:
            return None

    for rx in SCORE_REGEXES:
        m = rx.search(text)
        if m:
            try:
                val = float(m.group(1))
                unit = m.group(2) if m.lastindex and m.lastindex >= 2 else None
                pct = normalize_to_percent(val, unit)
                return clamp_pct(round(pct, 2)) if pct is not None else None
            except Exception:
                return None

    return None


def extract_score_openai_percent(client: OpenAI, text: str) -> Optional[float]:
    if not text:
        return None

    snippet = text[:MAX_CHARS_TO_MODEL]

    schema = {
        "name": "score_extraction_percent",
        "schema": {
            "type": "object",
            "additionalProperties": False,
            "properties": {
                "score_value": {"type": ["number", "null"]},
                "score_unit": {"type": "string", "enum": ["percent", "out_of_10", "unknown"]},
                "evidence": {"type": "string"},
            },
            "required": ["score_value", "score_unit", "evidence"],
        },
        "strict": True,
    }

    system_prompt = (
        "Extract the FINAL OVERALL SCORE from the given text.\n"
        "Examples:\n"
        "- 'Final Overall Score: 8/10'\n"
        "- 'Final Overall Score: 60%'\n"
        "- 'Final Overall Score: FAIL (30%)'\n"
        "Return score_value and score_unit.\n"
        "If no final score exists, score_value must be null.\n"
        "Do not guess."
    )

    try:
        resp = client.responses.create(
            model=DEFAULT_OPENAI_MODEL,
            input=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": snippet},
            ],
            text={"format": {"type": "json_schema", "json_schema": schema}},
            temperature=0,
        )
    except Exception:
        return None

    try:
        data = json.loads(safe_response_output_text(resp))
    except Exception:
        return None

    val = data.get("score_value", None)
    unit = data.get("score_unit", "unknown")
    if val is None:
        return None

    try:
        val = float(val)
    except Exception:
        return None

    unit_hint = "%" if unit == "percent" else ("/10" if unit == "out_of_10" else None)
    pct = normalize_to_percent(val, unit_hint)
    return clamp_pct(round(pct, 2)) if pct is not None else None


def pick_best_score_from_llm_outputs_percent(
    service,
    openai_client: OpenAI,
    files: List[dict],
) -> Tuple[Optional[float], Optional[str]]:
    llm_files = [
        f for f in files
        if f.get("mimeType") != FOLDER_MIME and LLM_OUTPUT_FILE_RE.match(f.get("name", ""))
    ]
    if not llm_files:
        return None, None

    llm_files.sort(key=lambda x: (x.get("modifiedTime") or ""), reverse=True)

    for f in llm_files:
        text = drive_download_text(service, f["id"], f.get("mimeType"))

        s = extract_score_regex_percent(text)
        if s is not None:
            return s, f.get("name")

        s = extract_score_openai_percent(openai_client, text)
        if s is not None:
            return s, f.get("name")

    return None, None


def autosize_columns(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            v = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(v))
        ws.column_dimensions[col_letter].width = min(max(12, max_len + 2), 60)


def set_percent_cell(cell, pct_value: Optional[float]):
    if pct_value is None:
        cell.value = None
        return
    cell.value = float(pct_value) / 100.0
    cell.number_format = "0%"
    cell.alignment = Alignment(horizontal="center", vertical="center")


def build_slot_workbook_percent(rows: List[Dict[str, Any]]) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Deliverables Analysis Sheet"

    headers = ["Person Name"] + FOLDER_NAMES + [AVG_HEADER]
    ws.append(headers)

    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    header_to_col = {h: i + 1 for i, h in enumerate(headers)}

    for r in rows:
        person = r["person"]
        ws.append([person] + [None] * (len(headers) - 1))
        row_idx = ws.max_row

        for folder_name in FOLDER_NAMES:
            col_idx = header_to_col[folder_name]
            pct = r.get(folder_name)
            cell = ws.cell(row=row_idx, column=col_idx)
            set_percent_cell(cell, pct)

            thr = THRESHOLDS_PCT.get(folder_name)
            if thr is not None and isinstance(pct, (int, float)) and float(pct) >= float(thr):
                cell.fill = GREEN_FILL

        scores_pct = [r.get(folder) for folder in FOLDER_NAMES]
        numeric = [s for s in scores_pct if isinstance(s, (int, float))]
        avg = round(sum(numeric) / len(numeric), 2) if numeric else None

        avg_col = header_to_col[AVG_HEADER]
        avg_cell = ws.cell(row=row_idx, column=avg_col)
        set_percent_cell(avg_cell, avg)

        thr_avg = THRESHOLDS_PCT.get(AVG_HEADER)
        if thr_avg is not None and isinstance(avg, (int, float)) and float(avg) >= float(thr_avg):
            avg_cell.fill = GREEN_FILL

        ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal="left", vertical="center")

    ws.freeze_panes = "A2"
    autosize_columns(ws)
    return wb


def main():
    if not SHARED_DRIVE_NAME:
        raise RuntimeError("SHARED_DRIVE_NAME is required. Example: SHARED_DRIVE_NAME=2026_Shared_Drive")

    openai_client = init_openai_client()
    service = get_drive_service()

    print("[INFO] Shared-Drive-only mode enabled")
    print(f"[INFO] SHARED_DRIVE_NAME={SHARED_DRIVE_NAME}")
    print(f"[INFO] ROOT_2026_FOLDER_NAME={ROOT_2026_FOLDER_NAME}")
    print(f"[INFO] OUTPUT_ROOT_FOLDER_NAME={OUTPUT_ROOT_FOLDER_NAME}")

    shared_drive = get_shared_drive_by_name(service, SHARED_DRIVE_NAME)
    shared_drive_id = shared_drive["id"]

    print(f"[INFO] Using Shared Drive: {shared_drive['name']} ({shared_drive_id})")

    if ROOT_2026_FOLDER_ID:
        print(f"[ROOT] Using ROOT_2026_FOLDER_ID={ROOT_2026_FOLDER_ID}")
        base_root = drive_get_file(service, ROOT_2026_FOLDER_ID)
        if base_root.get("mimeType") != FOLDER_MIME:
            raise RuntimeError(f"ROOT_2026_FOLDER_ID={ROOT_2026_FOLDER_ID} is not a folder.")
        file_drive_id = base_root.get("driveId")
        if file_drive_id and file_drive_id != shared_drive_id:
            raise RuntimeError(
                f"ROOT_2026_FOLDER_ID belongs to driveId={file_drive_id}, "
                f"but SHARED_DRIVE_NAME resolved to driveId={shared_drive_id}."
            )
    else:
        candidates_root = drive_search_folder_anywhere_in_shared_drive(service, ROOT_2026_FOLDER_NAME, shared_drive_id)
        if not candidates_root:
            raise RuntimeError(
                f"Could not find folder '{ROOT_2026_FOLDER_NAME}' inside Shared Drive '{SHARED_DRIVE_NAME}'."
            )
        if len(candidates_root) > 1:
            print(f"[WARN] Multiple folders named '{ROOT_2026_FOLDER_NAME}' inside Shared Drive '{SHARED_DRIVE_NAME}'.")
            for c in candidates_root[:10]:
                print(
                    " - id:", c["id"],
                    "modified:", c.get("modifiedTime"),
                    "parents:", c.get("parents"),
                    "driveId:", c.get("driveId"),
                )
            print("[WARN] Using the most recently modified one inside this Shared Drive.")
        base_root = pick_best_named_folder(candidates_root)
        print(f"[ROOT] Using source folder: {base_root['name']} (id={base_root['id']})")

    if OUTPUT_ROOT_FOLDER_ID:
        print(f"[ROOT] Using OUTPUT_ROOT_FOLDER_ID={OUTPUT_ROOT_FOLDER_ID}")
        output_root = drive_get_file(service, OUTPUT_ROOT_FOLDER_ID)
        if output_root.get("mimeType") != FOLDER_MIME:
            raise RuntimeError(f"OUTPUT_ROOT_FOLDER_ID={OUTPUT_ROOT_FOLDER_ID} is not a folder.")
        file_drive_id = output_root.get("driveId")
        if file_drive_id and file_drive_id != shared_drive_id:
            raise RuntimeError(
                f"OUTPUT_ROOT_FOLDER_ID belongs to driveId={file_drive_id}, "
                f"but SHARED_DRIVE_NAME resolved to driveId={shared_drive_id}."
            )
    else:
        candidates_out = drive_search_folder_anywhere_in_shared_drive(service, OUTPUT_ROOT_FOLDER_NAME, shared_drive_id)
        if not candidates_out:
            raise RuntimeError(
                f"Could not find output folder '{OUTPUT_ROOT_FOLDER_NAME}' inside Shared Drive '{SHARED_DRIVE_NAME}'. "
                f"Create it and run again."
            )
        if len(candidates_out) > 1:
            print(f"[WARN] Multiple folders named '{OUTPUT_ROOT_FOLDER_NAME}' inside Shared Drive '{SHARED_DRIVE_NAME}'.")
            for c in candidates_out[:10]:
                print(
                    " - id:", c["id"],
                    "modified:", c.get("modifiedTime"),
                    "parents:", c.get("parents"),
                    "driveId:", c.get("driveId"),
                )
            print("[WARN] Using the most recently modified one inside this Shared Drive.")
        output_root = pick_best_named_folder(candidates_out)
        print(f"[ROOT] Using output folder: {output_root['name']} (id={output_root['id']})")

    slot = choose_slot(service, base_root["id"], shared_drive_id)
    slot_name = slot["name"]
    slot_id = slot["id"]

    print(f"\n=== SLOT (SOURCE): {slot_name} ===")

    slot_out = drive_find_child(service, output_root["id"], slot_name, FOLDER_MIME, shared_drive_id)
    if not slot_out:
        slot_out_id = drive_create_folder(service, output_root["id"], slot_name)
        print(f"[OUTPUT] Created folder: {OUTPUT_ROOT_FOLDER_NAME}/{slot_name}")
    else:
        slot_out_id = slot_out["id"]

    print("[PERM] Setting folder editors...")
    drive_grant_editor_access(service, slot_out_id, EDITOR_EMAILS)

    people_all = sorted(
        list(drive_list_children(service, slot_id, FOLDER_MIME, shared_drive_id)),
        key=lambda x: (x.get("name") or "").lower(),
    )
    people = [p for p in people_all if (p.get("name") or "") not in SKIP_PERSON_FOLDERS]

    slot_rows: List[Dict[str, Any]] = []

    for person in people:
        person_name = person["name"]
        print(f" - Person: {person_name}")
        row: Dict[str, Any] = {"person": person_name}

        person_child_folders = list(drive_list_children(service, person["id"], FOLDER_MIME, shared_drive_id))
        folder_map = {
            f["name"]: f
            for f in person_child_folders
            if (f.get("name") or "") not in SKIP_DELIVERABLE_FOLDERS
        }

        for folder_name in FOLDER_NAMES:
            folder_node = folder_map.get(folder_name)
            if not folder_node:
                row[folder_name] = None
                print(f"   * {folder_name}: folder missing")
                continue

            files = list(drive_list_children(service, folder_node["id"], None, shared_drive_id))
            score_pct, used_filename = pick_best_score_from_llm_outputs_percent(service, openai_client, files)
            row[folder_name] = score_pct

            if score_pct is None:
                print(f"   * {folder_name}: no score found in any LLM_OUTPUT__* file")
            else:
                print(f"   * {folder_name}: {score_pct}% (from {used_filename})")

        slot_rows.append(row)

    wb = build_slot_workbook_percent(slot_rows)

    with tempfile.TemporaryDirectory() as td:
        td = Path(td)
        local_xlsx = td / OUTPUT_XLSX_NAME
        wb.save(local_xlsx)

        uploaded_file_id = drive_upload_xlsx(service, slot_out_id, OUTPUT_XLSX_NAME, local_xlsx, shared_drive_id)
        print(f"[OK] Uploaded: {OUTPUT_ROOT_FOLDER_NAME}/{slot_name}/{OUTPUT_XLSX_NAME}")

        print("[PERM] Setting sheet editors...")
        drive_grant_editor_access(service, uploaded_file_id, EDITOR_EMAILS)

    time.sleep(SLEEP_BETWEEN_UPLOADS_SEC)


if __name__ == "__main__":
    main()